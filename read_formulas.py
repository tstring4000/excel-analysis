# Filename: read_formulas.py
# Author: Tyler Stringer
# Date: 2024-09-03
# Details: Reads in formulas (from top n rows) from an Excel workbook.

import openpyxl
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def get_formulas(path, sheet_name, num_rows):
    logging.info(f"Opening workbook at {path}")
    try:
        workbook = openpyxl.load_workbook(path, data_only=False)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

        logging.info(f"Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        formulas = set()

        for row in sheet.iter_rows(min_row=1, max_row=num_rows):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.add("'" + cell.value)  # Prepend apostrophe to treat as text

        if not formulas:
            logging.info(f"No formulas found in this sheet: {sheet_name}")
        else:
            formulas_list = list(formulas)  # Convert set to list to log the first few elements
            logging.info(f"Found formulas: {formulas_list[1:3]} ...")
        return formulas

    except FileNotFoundError:
        logging.error(f"The file at '{path}' was not found.")
        return set()
    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")
        return set()


if __name__ == '__main__':
    path = 'data/input/FILENAME.xlsx'
    num_rows = 3

    # Extract unique formulas from the specified number of rows of the first sheet
    sheet_names = ['Sheet1']  # Replace with dynamic sheet name extraction if needed
    first_sheet_name = sheet_names[0]
    formulas = get_formulas(path, first_sheet_name, num_rows)
    print(f"Formulas in '{first_sheet_name}':", formulas)
